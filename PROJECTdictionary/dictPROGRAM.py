import openpyxl as p
import os, sys
print("WELCOME TO ADMIN ACCOUNT!!!")
if(input('PLEASE ENTER YOUR PASSWORD :- ')=='Rithvik'):
    d=p.load_workbook(os.path.split(os.path.abspath(os.path.realpath(sys.argv[0])))[0]+'\\DICTIONARY.xlsx')
    s1=d["Sheet1"]
    count=[]
    def Move(row):
        s1.insert_rows(count[row])
        s1.cell(count[row],1).value=word
        s1.cell(count[row],2).value=mean
        for i in range(row,26):
            count[i]=count[i]+1
        for i in range(3,29):
            s1.cell(2,i).value=count[i-3]    
        d.save("I:\\rithvikPython\\PROJECTdictionary\\DICTIONARY.xlsx")    
    for i in range(3,29):
        count.append(s1.cell(2,i).value)
    while(True):
        word=input("enter the WORD :- ").lower()
        mean=input("enter the MEANING :- ").lower()
        a=ord(word[0])
        if(a>=97 and a<=122):
            Move(a-97)
        if(str(input('you wanna enter more??')).lower()=="no"):
           break
    d.save("I:\\rithvikPython\\PROJECTdictionary\\DICTIONARY.xlsx")
else:
    print('Sorry you are not eligible for accseing this APPLICATION')
input('enter any key to exit')
