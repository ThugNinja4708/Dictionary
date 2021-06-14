import openpyxl as p
import os,sys
print("WELCOME TO USER ACCOUNT!!!")
d=p.load_workbook(os.path.split(os.path.abspath(os.path.realpath(sys.argv[0])))[0]+'\\DICTIONARY.xlsx')
s1=d["Sheet1"]
n=3
count=[]
for i in range(3,29):
    count.append(s1.cell(2,i).value)
def fun():
    word=input('ENTER THE SERCH ELEMENT :- ').lower()
    a=ord(word[0])-97
    if(a==0):
        serch=s1.cell(n,1).value
        while(serch!=None):
            serch=s1.cell(n,1).value
            n+=1
            if(serch==word):
                print('********THE SERCH ELEMENT IS FOUND!!********')
                print(serch,' = ',s1.cell(n,2).value)
        print('********NOT FOUND********')
    elif(a>0):
        n=count[a-1]+2
        serch=s1.cell(n,1).value
        while(serch!=None):
            if(serch==word):
                print('********THE SERCH ELEMENT IS FOUND!!********')
                print(serch,' = ',s1.cell(n,2).value)
            n+=1
            serch=s1.cell(n,1).value
    else:
        print('********NOT FOUND********')
while(True):
    fun()
    if(input('Wanna serch more??(y/n)').lower()=='n'):
        input('Enter any key to exit')
        exit(0)
        
    
